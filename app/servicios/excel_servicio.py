from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def generar_excel_alumnos(curso_titulo, estudiantes_data):
    """
    Genera un archivo Excel profesional con la información de los alumnos.

    :param curso_titulo: Nombre del curso
    :param estudiantes_data: Lista de dicts con claves:
        - numero_control, nombre, email, carrera, semestre, grupo,
          progreso, completadas, total_lecciones, avg_puntaje, estado
    :returns: BytesIO buffer con el archivo .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Alumnos"

    # ─── Estilos ───
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_font = Font(name='Calibri', bold=True, size=14, color='003366')
    subtitle_font = Font(name='Calibri', size=10, color='666666')

    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')

    alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    green_font = Font(name='Calibri', size=10, color='16A34A', bold=True)
    orange_font = Font(name='Calibri', size=10, color='D97706', bold=True)
    gray_font = Font(name='Calibri', size=10, color='94A3B8')

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # ─── Título ───
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f'📋 Reporte de Alumnos — {curso_titulo}'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A2:I2')
    sub_cell = ws['A2']
    sub_cell.value = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")} · Matatucas LMS'
    sub_cell.font = subtitle_font

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # ─── Encabezados (fila 4) ───
    headers = [
        'No. Control',
        'Nombre Completo',
        'Correo',
        'Carrera',
        'Semestre',
        'Grupo',
        'Progreso (%)',
        'Lecciones',
        'Prom. Ejercicios',
        'Estado',
    ]

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[header_row].height = 28

    # ─── Datos ───
    for row_idx, est in enumerate(estudiantes_data, start=header_row + 1):
        row_data = [
            est.get('numero_control', '-'),
            est.get('nombre', ''),
            est.get('email', ''),
            est.get('carrera', 'Sin especificar'),
            est.get('semestre', '-'),
            est.get('grupo', '-'),
            est.get('progreso', 0),
            f"{est.get('completadas', 0)} / {est.get('total_lecciones', 0)}",
            round(est.get('avg_puntaje', 0), 1),
            est.get('estado', '-'),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = center_alignment if col_idx >= 5 else data_alignment
            cell.border = thin_border

            # Alternar color de fila
            if (row_idx - header_row) % 2 == 0:
                cell.fill = alt_fill

        # Colorear estado
        estado_cell = ws.cell(row=row_idx, column=10)
        estado_val = str(estado_cell.value).lower()
        if 'completado' in estado_val:
            estado_cell.font = green_font
        elif 'progreso' in estado_val:
            estado_cell.font = orange_font
        else:
            estado_cell.font = gray_font

    # ─── Anchos automáticos ───
    column_widths = [16, 30, 30, 25, 10, 10, 14, 14, 16, 14]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ─── Congelar encabezado ───
    ws.freeze_panes = f'A{header_row + 1}'

    # ─── Guardar en buffer ───
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
